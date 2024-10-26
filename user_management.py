import openpyxl
import os

filename="user_data.xlsx"

def add_user():
    name=input("Enter name: ")
    email=input("Enter email: ")
    phone=input("Enter phone number: ")

    if os.path.exists(filename):
        workbook=openpyxl.load_workbook(filename)
        sheet=workbook.active
    else:
        workbook=openpyxl.Workbook()
        sheet=workbook.active
        sheet.append(["Name","Email","Phone"])

    sheet.append([name,email, phone])
    workbook.save(filename)
    print("User added succesfully!!")

def display_users():
    if os.path.exists(filename):
        workbook=openpyxl.load_workbook(filename)
        sheet=workbook.active

        print("\nStored Users:\n")
        print(f"{'Name':<20}{'Email':<30}{'Phone':<15}")
        print("-"*65)
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]:<20}{row[1]:<30}{row[2]:<15}")
    else:
        print("No user found. Add user first!")

def main():
    while True:
        print("\nMake a choice to proceed: ")
        print("Press '1' to add user")
        print("Press '2' to display users")
        print("Press '3' to exit")

        choice = input("\nPress any key to countinue: ")

        if choice == '1':
            add_user()
        elif choice == '2':
            display_users()
        elif choice == '3':
            print("Exiting program, Goodbye!!")
            break
        else:
            print("Invalid choice!")

if __name__ == '__main__':
    main()


